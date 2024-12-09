import os
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib.auth.models import User, Group
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
import pandas as pd
from django.core.exceptions import MultipleObjectsReturned
from .forms import PatientRecordForm, LabTestResultForm, PrescriptionForm
from .models import PatientRecord
from django.urls import reverse
from .forms import PatientForm  
from datetime import datetime
from django.contrib.auth import update_session_auth_hash
from django.utils.timezone import localtime
from django.core.files.storage import FileSystemStorage
from .forms import PDFUploadForm
from .models import ExtractedData
from .utils import extract_content_from_pdf  
from django.views.decorators.csrf import csrf_exempt
from .models import PatientRecord  
from django.contrib import messages
from .forms import UploadFileForm
from .models import Employee
from openpyxl import load_workbook

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)  
            return redirect_user_based_on_role(request, user)  
        else:
            return render(request, 'login.html', {'error': 'Invalid username or password.'})

    return render(request, 'login.html')

def redirect_user_based_on_role(request, user):
    if user.groups.filter(name='super admin').exists():
        return redirect(reverse('admin:index'))  
    elif user.groups.filter(name='admin').exists():
        return redirect(reverse('admin:index'))  
    elif user.groups.filter(name='doctor').exists():
        return redirect('doctor_dashboard')  
    elif user.groups.filter(name='lab technician').exists():
        return redirect('lab_technician_dashboard')  
    elif user.groups.filter(name='human resource').exists():
        return redirect('human_resource_report')
    return render(request, 'login.html', {'error': 'User has no assigned role.'})   


@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

def check_user_group(user, group_name):
    return user.groups.filter(name=group_name).exists()

@login_required
@user_passes_test(lambda u: check_user_group(u, 'super admin'))
def admin_dashboard(request):
    return redirect(reverse('admin:index'))

@login_required
@user_passes_test(lambda u: check_user_group(u, 'admin'))
def admin_dashboard(request):
    return redirect(reverse('admin:index'))

@login_required
@user_passes_test(lambda u: check_user_group(u, 'doctor'))
def doctor_dashboard(request):
    return handle_form_submission(request, PatientRecordForm, 'doctor_dashboard.html', 'patient_history')

@login_required
@user_passes_test(lambda u: check_user_group(u, 'lab technician'))
def lab_technician_dashboard(request):
    return render(request, 'lab_technician_dashboard.html')

@login_required
@user_passes_test(lambda u: check_user_group(u, 'human resource'))
def human_resource_report(request):
    return render(request, 'human_resource_report.html')

@login_required
@user_passes_test(lambda u: check_user_group(u, 'super admin'))
def add_user(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        role = request.POST['role']
        user = User.objects.create_user(username=username, password=password)
        group = Group.objects.get(name=role)
        user.groups.add(group)
        return redirect(reverse('admin:index'))
    return redirect(reverse('admin:index'))

@login_required
@user_passes_test(lambda u: check_user_group(u, 'super admin'))
def manage_users(request):
    users = User.objects.all()
    return render(request, 'manage_users.html', {'users': users})

@login_required
def check_unique_code(request):
    if request.method == 'POST':
        unique_code = request.POST.get('unique_code')
        try:
            patient = PatientRecord.objects.get(unique_code=unique_code)
            response_data = {
                'test_type': patient.test_type,
                'description': patient.description,
                'valid': True
            }
        except PatientRecord.DoesNotExist:
            response_data = {
                'error': 'No patient found with that unique code.',
                'valid': False
            }
        except MultipleObjectsReturned:
            response_data = {
                'error': 'Multiple patients found with that unique code. Please check for duplicates.',
                'valid': False
            }
        return JsonResponse(response_data)

@login_required
def lab_technician_view(request):
    form = LabTestResultForm()
    test_type = description = None

    if request.method == 'POST':
        unique_code = request.POST.get('unique_code')
        try:
            patient = PatientRecord.objects.get(unique_code=unique_code)
            test_type = patient.test_type
            description = patient.description
            form = LabTestResultForm(request.POST, instance=patient)
            if form.is_valid():
                form.save()
                messages.success(request, 'Lab test result submitted successfully.')
                return redirect('lab_reports')
            else:
                messages.error(request, 'Please correct the errors below.')
        except PatientRecord.DoesNotExist:
            messages.error(request, 'No patient found with that unique code.')
        except MultipleObjectsReturned:
            messages.error(request, 'Multiple patients found with that unique code. Please check for duplicates.')

    return render(request, 'lab_technician.html', {
        'form': form,
        'test_type': test_type,
        'description': description,
        'messages': messages.get_messages(request),
    })

@login_required
def add_prescription(request, patient_id):
    patient = get_object_or_404(PatientRecord, id=patient_id)
    if not patient.test_result:
        return redirect('patient_history')
    return handle_form_submission(request, PrescriptionForm, 'add_prescription.html', 'patient_history', patient_id)

@login_required
def get_name_by_clock_number(request):
    clock_number = request.GET.get('clock_number')
    name = "N/A"
    
    if clock_number:
        try:
            patient_record = PatientRecord.objects.get(clock_number=clock_number)
            name = patient_record.name
        except PatientRecord.DoesNotExist:
            pass
    
    return JsonResponse({'name': name})
@login_required
def patient_history_view(request):
    return redirect_user_based_on_role(request.user) 

@login_required
def patient_history(request):
    search_query = request.GET.get('search_query', '').strip()
    gender_filter = request.GET.get('gender_filter', '').strip()
    month_filter = request.GET.get('month_filter', '').strip()
    start_date = request.GET.get('start_date', '').strip()  
    end_date = request.GET.get('end_date', '').strip()     

    records = PatientRecord.objects.all().order_by('-created_at')

    if search_query:
        records = records.filter(
            Q(clock_number__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(unique_code__icontains=search_query)
        )

    if gender_filter:
        records = records.filter(gender=gender_filter)

    if month_filter:
        try:
            year, month = map(int, month_filter.split('/'))
            records = records.filter(created_at__year=year, created_at__month=month)
        except ValueError:
            print("Error parsing the month filter")

    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')

            records = records.filter(created_at__range=(start_date_obj, end_date_obj))

            print(f"Filtering from: {start_date_obj} to {end_date_obj}")
            print(f"Records found after filtering: {records.count()}")

        except ValueError:
            print("Error parsing the date range filters")

    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    records = paginator.get_page(page_number)

    no_records_found = not records

    return render(request, 'patient_history.html', {'records': records, 'no_records_found': no_records_found})

@login_required
def update_patient_record(request, unique_code):
    patient = get_object_or_404(PatientRecord, unique_code=unique_code)   
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)  
        if form.is_valid():
            form.save()   
            return redirect('patient_history')   
    else:
        form = PatientForm(instance=patient)   
    return render(request, 'update_patient_record.html', {'form': form, 'patient': patient})

@login_required
def create_patient_record(request):
    return handle_form_submission(request, PatientRecordForm, 'doctor_dashboard.html', 'patient_history')

@login_required
def lab_update_record(request, unique_code):
    patient = get_object_or_404(PatientRecord, unique_code=unique_code)
    
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            return redirect('lab_update_record', unique_code=unique_code)
        else:
            print(form.errors)  
    
    else:
        form = PatientForm(instance=patient)   
    
    return render(request, 'lab_update_record.html', {'patient': patient, 'form': form})


@login_required
def lab_reports(request):
    search_query = request.GET.get('search_query', '').strip()
    gender_filter = request.GET.get('gender_filter', '').strip()
    month_filter = request.GET.get('month_filter', '').strip() 

    records = PatientRecord.objects.all().order_by('-created_at')  

    if search_query:
        records = records.filter(
            Q(clock_number__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(unique_code__icontains=search_query)
        )

    if gender_filter:
        records = records.filter(gender=gender_filter)

    if month_filter:
        try:
            year, month = map(int, month_filter.split('/')) 
            records = records.filter(created_at__year=year, created_at__month=month)
        except ValueError:
            print(f"Invalid month filter: {month_filter}. Expected format: 'YYYY/MM'.")

    paginator = Paginator(records, 20) 
    page_number = request.GET.get('page')
    records = paginator.get_page(page_number)  

    no_records_found = not records 

    return render(request, 'lab_report.html', {
        'records': records,
        'no_records_found': no_records_found,
        'search_query': search_query,
        'gender_filter': gender_filter,
        'month_filter': month_filter
    })


@login_required
def human_resource_report(request):
    search_query = request.GET.get('search_query', '').strip()
    gender_filter = request.GET.get('gender_filter', '').strip()
    month_filter = request.GET.get('month_filter', '').strip() 

    records = PatientRecord.objects.all().order_by('-created_at')  

    if search_query:
        records = records.filter(
            Q(clock_number__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(unique_code__icontains=search_query)
        )

    if gender_filter:
        records = records.filter(gender=gender_filter)

    if month_filter:
        try:
            year, month = map(int, month_filter.split('/')) 
            records = records.filter(created_at__year=year, created_at__month=month)
        except ValueError:
            print(f"Invalid month filter: {month_filter}. Expected format: 'YYYY/MM'.")

    paginator = Paginator(records, 20) 
    page_number = request.GET.get('page')
    records = paginator.get_page(page_number)  

    no_records_found = not records 

    return render(request, 'human_resource_report.html', {
        'records': records,
        'no_records_found': no_records_found,
        'search_query': search_query,
        'gender_filter': gender_filter,
        'month_filter': month_filter
    })


@login_required
def export_to_excel(request):
    print(f'Request Method: {request.method}')
    print(f'GET Parameters: {request.GET}')
    print(f'POST Parameters: {request.POST}')

    search_query = request.POST.get('search_query', '').strip()
    gender_filter = request.POST.get('gender_filter', '').strip()
    month_filter = request.POST.get('month_filter', '').strip()
    start_date = request.POST.get('start_date', '').strip()
    end_date = request.POST.get('end_date', '').strip()

    selected_records = request.POST.getlist('selected_records')  

    records = PatientRecord.objects.all()

    print(f'Gender Filter: {gender_filter}')
    print(f'Search Query: {search_query}')
    print(f'Selected Records: {selected_records}')
    print(f'Initial Records Count: {records.count()}')

    if selected_records:
        records = records.filter(unique_code__in=selected_records)
        print(f'Filtered by Selected Records Count: {records.count()}')

    if search_query:
        records = records.filter(
            Q(clock_number__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(unique_code__icontains=search_query)
        )
        print(f'After Search Filter Count: {records.count()}')

    if gender_filter:
        records = records.filter(gender=gender_filter)
        print(f'After Gender Filter Count: {records.count()}')

    if month_filter:
        try:
            year, month = map(int, month_filter.split('/'))
            records = records.filter(created_at__year=year, created_at__month=month)
            print(f'After Month Filter Count: {records.count()}')
        except ValueError:
            print("Error parsing the month filter")

    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            records = records.filter(created_at__range=(start_date_obj, end_date_obj))
            print(f'After Date Range Filter Count: {records.count()}')
        except ValueError:
            print("Error parsing the date range filters")

    if not records.exists():
        return HttpResponse("No records found based on the filters applied.")

    print(f'Final Records Count for Export: {records.count()}')

    data = {
        'Clock Number': [record.clock_number for record in records],
        'Name': [record.name for record in records],
        'Gender': [record.gender for record in records],
        'age': [record.age for record in records],
        'vitals':[record.vitals for record in records],
        'Next Of Kin': [record.relationship for record in records],
        'chief compliants': [record.chief_complaints for record in records],
        'History of Presenting Illness': [record.history_of_presenting_illness for record in records],
        'Date Symptoms Aware': [record.date_symptoms_aware.strftime('%Y-%m-%d') if record.date_symptoms_aware else '' for record in records],
        'Previous Complaints': [record.previous_complaints for record in records],
        'Treatment Description': [record.treatment_description for record in records],
        'medical history': [record.medical_history for record in records],
        'Examination': [record.examination for record in records],
        'Test Type': [record.test_type for record in records],
        'Description': [record.description for record in records],
        'Test Result': [record.test_result for record in records],
        'Diagnosis': [record.diagnosis for record in records],
        'Treatment': [record.prescription for record in records],
        'medical remarks':[record.medical_remarks for record in records],
        'Remarks': [record.remarks for record in records],
        'Date Created': [localtime(record.created_at).strftime('%Y-%m-%d %H:%M:%S') if record.created_at else '' for record in records],
    }

    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="selected_patient_history.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    return response

@login_required
def hr_export_to_excel(request):
    print(f'Request Method: {request.method}')
    print(f'GET Parameters: {request.GET}')
    print(f'POST Parameters: {request.POST}')

    search_query = request.POST.get('search_query', '').strip()
    gender_filter = request.POST.get('gender_filter', '').strip()
    month_filter = request.POST.get('month_filter', '').strip()
    start_date = request.POST.get('start_date', '').strip()
    end_date = request.POST.get('end_date', '').strip()

    selected_records = request.POST.getlist('selected_records')  

    records = PatientRecord.objects.all()

    print(f'Gender Filter: {gender_filter}')
    print(f'Search Query: {search_query}')
    print(f'Selected Records: {selected_records}')
    print(f'Initial Records Count: {records.count()}')

    if selected_records:
        records = records.filter(unique_code__in=selected_records)
        print(f'Filtered by Selected Records Count: {records.count()}')

    if search_query:
        records = records.filter(
            Q(clock_number__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(unique_code__icontains=search_query)
        )
        print(f'After Search Filter Count: {records.count()}')

    if gender_filter:
        records = records.filter(gender=gender_filter)
        print(f'After Gender Filter Count: {records.count()}')

    if month_filter:
        try:
            year, month = map(int, month_filter.split('/'))
            records = records.filter(created_at__year=year, created_at__month=month)
            print(f'After Month Filter Count: {records.count()}')
        except ValueError:
            print("Error parsing the month filter")

    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            records = records.filter(created_at__range=(start_date_obj, end_date_obj))
            print(f'After Date Range Filter Count: {records.count()}')
        except ValueError:
            print("Error parsing the date range filters")

    if not records.exists():
        return HttpResponse("No records found based on the filters applied.")

    print(f'Final Records Count for Export: {records.count()}')

    data = {
        'Clock Number': [record.clock_number for record in records],
        'Name': [record.name for record in records],
        'Gender': [record.gender for record in records],
        'age': [record.age for record in records],
        'Next Of Kin': [record.relationship for record in records],     
        'Remarks': [record.remarks for record in records],
        'Date Created': [localtime(record.created_at).strftime('%Y-%m-%d %H:%M:%S') if record.created_at else '' for record in records],
    }

    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="selected_patient_history.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    return response

@login_required
def patient_records_by_clock_number(request, clock_number):
    records = PatientRecord.objects.filter(clock_number=clock_number).order_by('-created_at')
    return render(request, 'patient_records_by_clock_number.html', {'records': records})

@login_required
def handle_form_submission(request, form_class, template, success_redirect, instance_id=None):
    instance = get_object_or_404(PatientRecord, id=instance_id) if instance_id else None

    if request.method == 'POST':
        form = form_class(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            return redirect(success_redirect)
    else:
        form = form_class(instance=instance)

    return render(request, template, {'form': form, 'instance': instance})

@login_required
def doctor_dashboard(request):
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            record = form.save()
            messages.success(request, 'Patient record created successfully.')
            return redirect('doctor_dashboard')
    else:
        form = PatientForm()

    records = PatientRecord.objects.all()

    sort_by = request.GET.get('sort_by', 'created_at')
    sort_order = request.GET.get('sort_order', 'asc') 
    if sort_order == 'desc':
        records = records.order_by(f'-{sort_by}')
    else:
        records = records.order_by(sort_by)

    return render(request, 'doctor_dashboard.html', {
        'form': form,
        'records': records,
        'sort_by': sort_by,  
        'sort_order': sort_order  
    })

@login_required
def get_patient_details(request):
    clock_number = request.GET.get('clock_number')
    if clock_number:
        try:
            patient = PatientRecord.objects.get(clock_number=clock_number)
            data = {
                'name': patient.name,
                'dob': patient.dob.strftime('%Y-%m-%d'),
            }
        except PatientRecord.DoesNotExist:
            data = {'error': 'Patient not found'}
    else:
        data = {'error': 'Clock number not provided'}
    
    return JsonResponse(data)

@login_required
def auto_logout(request):
    logout(request)
    return JsonResponse({'status': 'logged_out'})

@login_required
def renew_session(request):
    if request.user.is_authenticated:
        request.session.set_expiry(1800)
        return JsonResponse({'status': 'session_renewed'})
    return JsonResponse({'status': 'not_authenticated'})

@login_required
def submit_test_info(request):
    if request.method == 'POST':
        form = PatientRecordForm(request.POST, request.FILES)  
        if form.is_valid():
            patient_record = form.save(commit=False)

           
            if 'test_result_pdf' in request.FILES:
                uploaded_file = request.FILES['test_result_pdf']


                fs = FileSystemStorage()
                filename = fs.save(uploaded_file.name, uploaded_file)
                patient_record.test_result_pdf = fs.url(filename)

            patient_record.save()
            print('PatientRecord')
            messages.success(request, "Patient record submitted successfully!") 
            return redirect('lab_report') 

        else:
            messages.error(request, "There was an error submitting the form. Please try again.") 

    else:
        form = PatientRecordForm()

    return render(request, 'lab_technician.html', {'form': form})

@login_required
def upload_pdf_view(request):
    if request.method == 'POST':
        form = PDFUploadForm(request.POST, request.FILES)
        if form.is_valid():
            pdf_file = request.FILES['pdf_file']
            file_path = os.path.join('uploads', pdf_file.name)

            if not os.path.exists('uploads'):
                os.makedirs('uploads')

            with open(file_path, 'wb+') as destination:
                for chunk in pdf_file.chunks():
                    destination.write(chunk)

            extracted_content, _ = extract_content_from_pdf(file_path)

            formatted_content = extracted_content.replace('\n', '<br>')

            extracted_data = ExtractedData(content=formatted_content)
            extracted_data.save()
            return render(request, 'lab_success.html', {'extracted_content': formatted_content})
    else:
        form = PDFUploadForm()

    return render(request, 'upload_pdf.html', {'form': form})



@login_required
def success_view(request):
    return render(request, 'upload_pdf.html')

@login_required
def save_patient_record(request):
    if request.method == 'POST':
        form = PatientRecordForm(request.POST)
        if form.is_valid():
            patient_record = form.save(commit=False)
            extracted_content = request.POST.get('extracted_content')

            patient_record.save() 
            return redirect('lab_report')  
    else:
        form = PatientRecordForm()

    return render(request, 'lab_report.html', {'form': form})

@login_required
@csrf_exempt  
def upload_pdf(request):
    if request.method == "POST":
        unique_code = request.POST.get('unique_code')
        pdf_file = request.FILES.get('test_result_pdf')
        
        if pdf_file:
            try:
                patient_record = PatientRecord.objects.filter(unique_code=unique_code).first()

                if patient_record is None:
                    return JsonResponse({'success': False, 'error': 'Patient record not found.'})

                patient_record.test_result_pdf = pdf_file
                patient_record.save() 

                return JsonResponse({'success': True, 'message': 'PDF uploaded successfully.'})

            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
        else:
            return JsonResponse({'success': False, 'error': 'No PDF file was uploaded.'})

    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

@login_required
def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            if file.name.endswith('.xlsx'):
                try:
                    workbook = load_workbook(file)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 5:
                            payroll_number, name, department, designation, status, *_ = row  

                            
                            if isinstance(payroll_number, float):
                                payroll_number = int(payroll_number)

                          
                            if isinstance(status, (int, float)):
                                status = bool(int(status))  
                            else:
                                status = True  

                            if isinstance(payroll_number, int) and isinstance(name, str):
                                Employee.objects.update_or_create(
                                    payroll_number=payroll_number,
                                    defaults={
                                        'name': name,
                                        'department': department if isinstance(department, str) else None,
                                        'designation': designation if isinstance(designation, str) else None,
                                        'status': status
                                    }
                                )
                            else:
                                messages.warning(request, f"Invalid data for Payroll No: {payroll_number}. Skipping this record.")
                        else:
                            messages.warning(request, f"Row has insufficient data. Skipping: {row}.")
                    messages.success(request, 'All records have been uploaded successfully!')
                except Exception as e:
                    messages.error(request, f"Error processing file: {str(e)}")
            else:
                messages.error(request, 'Invalid file type. Please upload an .xlsx file.')
    else:
        form = UploadFileForm()

    return render(request, 'upload.html', {'form': form})



@login_required
def get_employee(request, clock_number):
    try:
        employee = Employee.objects.get(payroll_number=clock_number, status=True)
        return JsonResponse({'name': employee.name})
    except Employee.DoesNotExist:
        return JsonResponse({'name': None})
    except Exception as e:
        return JsonResponse({'error': str(e)})


@login_required
def get_employee_data(request):
    payroll_number = request.GET.get('payroll_number')
    try:
        employee = Employee.objects.get(payroll_number=payroll_number)
        return JsonResponse({
            'department': employee.department,
            'designation': employee.designation
        })
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)

