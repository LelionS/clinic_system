from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.models import User, Group
from django.apps import apps
from django.db.models.signals import post_migrate

def create_user_groups(sender, **kwargs):
    roles = ['admin', 'doctor', 'lab technician']
    for role in roles:
        Group.objects.get_or_create(name=role)

post_migrate.connect(create_user_groups, sender=apps.get_app_config('doctors'))

class CustomUserAdmin(UserAdmin):
    list_display = ('username', 'email', 'first_name', 'last_name', 'is_staff')
    list_filter = ('is_staff', 'groups')
    fieldsets = (
        (None, {'fields': ('username', 'password')}),
        ('Personal info', {'fields': ('first_name', 'last_name', 'email')}),
        ('Permissions', {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups')}),
    )


from django.contrib import admin
from django.urls import path
from django.shortcuts import render
from django.contrib import messages
from openpyxl import load_workbook
from .models import Employee
from .forms import UploadFileForm

@admin.register(Employee)
class EmployeeAdmin(admin.ModelAdmin):
    list_display = ('payroll_number', 'name', 'department', 'designation', 'status')  # Show more fields

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('employee/', self.admin_site.admin_view(self.upload_view), name='employee-upload'),
        ]
        return custom_urls + urls

    def upload_view(self, request):
        if request.method == 'POST':
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES['file']
                if file.name.endswith('.xlsx'):
                    try:
                        workbook = load_workbook(file)
                        sheet = workbook.active
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            # Get the first five columns explicitly
                            payroll_number = row[0]
                            name = row[1]
                            department = row[2] if len(row) > 2 else None
                            designation = row[3] if len(row) > 3 else None
                            status = row[4] if len(row) > 4 else None

                            if isinstance(payroll_number, float):
                                payroll_number = int(payroll_number)  # Convert float to int if needed

                            # Convert status from 0 or 1 to boolean
                            if isinstance(status, (int, float)):
                                status = bool(int(status))  # Converts 0 to False and 1 to True
                            else:
                                status = True  # Default to True if status is not an integer

                            # Validate fields before creating/updating
                            if isinstance(payroll_number, int) and isinstance(name, str):
                                Employee.objects.update_or_create(
                                    payroll_number=payroll_number,
                                    defaults={
                                        'name': name,
                                        'department': department if isinstance(department, str) else None,
                                        'designation': designation if isinstance(designation, str) else None,
                                        'status': status
                                    }
                                )
                            else:
                                messages.warning(request, f"Invalid data for Payroll No: {payroll_number}. Skipping this record.")
                        messages.success(request, 'All records have been uploaded successfully!')
                    except Exception as e:
                        messages.error(request, f"Error processing file: {str(e)}")
                else:
                    messages.error(request, 'Invalid file type. Please upload an .xlsx file.')
        else:
            form = UploadFileForm()

        context = {
            'form': form,
            'title': 'Upload Employee Data',
        }
        return render(request, 'upload.html', context)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['upload_form'] = UploadFileForm()
        return super().changelist_view(request, extra_context=extra_context)

admin.site.unregister(User)
admin.site.register(User, CustomUserAdmin)


admin.site.site_header = 'Fides Clinic'
admin.site.site_title = 'My Admin Title'
admin.site.index_title = 'Admin!'